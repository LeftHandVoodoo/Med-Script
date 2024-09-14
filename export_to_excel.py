import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

def export_medications_to_excel(medications, filename='medications.xlsx'):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Medications"

    # Define headers
    headers = ['ID', 'Name', 'Strength', 'Dosage Frequency', 'Description']
    
    # Style for headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Write headers
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Write data
    for row, med in enumerate(medications, start=2):
        ws.cell(row=row, column=1, value=med['id'])
        ws.cell(row=row, column=2, value=med['name'])
        ws.cell(row=row, column=3, value=med['strength'])
        ws.cell(row=row, column=4, value=med['dosage_frequency'])
        ws.cell(row=row, column=5, value=med['description'])

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Save the workbook
    wb.save(filename)

if __name__ == "__main__":
    # Test data
    medications = [
        {'id': 1, 'name': 'Aspirin', 'strength': '81mg', 'dosage_frequency': 'Once daily', 'description': 'Pain relief, fever reduction'},
        {'id': 2, 'name': 'Lisinopril', 'strength': '10mg', 'dosage_frequency': 'Once daily', 'description': 'High blood pressure treatment'},
    ]
    export_medications_to_excel(medications)
    print("Excel file created successfully.")
